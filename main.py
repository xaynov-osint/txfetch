"""
txfetch — entry point.
tui on prompt_toolkit: ascii art → network selection → step-by-step input.
"""

import sys
import os
import argparse
from datetime import datetime, timezone, timedelta

from prompt_toolkit import Application
from prompt_toolkit.key_binding import KeyBindings
from prompt_toolkit.layout import Layout
from prompt_toolkit.layout.containers import HSplit, Window
from prompt_toolkit.layout.controls import FormattedTextControl
from prompt_toolkit.styles import Style
from prompt_toolkit.shortcuts import prompt, clear

ASCII_ART = """\
 ████████╗██╗  ██╗███████╗███████╗████████╗ ██████╗██╗  ██╗
 ╚══██╔══╝╚██╗██╔╝██╔════╝██╔════╝╚══██╔══╝██╔════╝██║  ██║
    ██║    ╚███╔╝ █████╗  █████╗     ██║   ██║     ███████║
    ██║    ██╔██╗ ██╔══╝  ██╔══╝     ██║   ██║     ██╔══██║
    ██║   ██╔╝ ██╗██║     ███████╗   ██║   ╚██████╗██║  ██║
    ╚═╝   ╚═╝  ╚═╝╚═╝     ╚══════╝   ╚═╝    ╚═════╝╚═╝  ╚═╝"""

SUBTITLE = "  Blockchain Transaction Retrieval Tool  ·  by xaynov (@ownstates)"

NETWORKS = [
    ("TRON",  "USDT · TRC-20"),
    ("TON",   "TON · Jetton"),
    ("ETH",   "EVM · ERC-20"),
    ("BTC",   "Bitcoin"),
]

# time range options shown after entering a timestamp
TIME_RANGES = [
    ("Exact minute",   "from HH:MM:00 to HH:MM:59",  "minute"),
    ("Exact second",   "only the specified second",    "second"),
    ("± 1 minute",     "30 seconds each side",         "1min"),
    ("± 5 minutes",    "5 minutes each side",          "5min"),
    ("± 15 minutes",   "15 minutes each side",         "15min"),
    ("Custom range",   "enter start and end manually", "custom"),
]

STYLE = Style.from_dict({
    "art":       "#5a9f7a bold",
    "subtitle":  "#444444",
    "divider":   "#2a2a2a",
    "label":     "#444444",
    "cursor":    "#5a9f7a bold",
    "item":      "#777777",
    "item.sel":  "#d0d0d0 bold",
    "hint-key":  "#5a9f7a",
    "hint":      "#333333",
    "tag":       "#333333",
})

COINS = {
    "tron": ["USDT"],
    "ton":  ["TON", "USDT", "NOT"],
    "eth":  ["USDT", "ETH", "USDC"],
    "bsc":  ["USDT", "BNB", "USDC"],
    "base": ["USDC", "ETH"],
    "btc":  ["BTC"],
}

EVM_NETWORKS = [
    ("ETH",  "Ethereum mainnet"),
    ("BSC",  "BNB Smart Chain"),
    ("BASE", "Base (Coinbase L2)"),
]

EXCEL_THRESHOLD = 100

def build_menu(selected: int, items: list, title: str = "SELECT NETWORK", show_art: bool = True):
    lines = []
    if show_art:
        lines.append(("class:art",      ASCII_ART + "\n"))
        lines.append(("class:subtitle", SUBTITLE + "\n"))
        lines.append(("",               "\n"))
    lines.append(("class:divider",  "─" * 62 + "\n"))
    lines.append(("class:label",    f"  {title}\n\n"))

    for i, (name, tag) in enumerate(items):
        if i == selected:
            lines += [("class:cursor",  "  › "),
                      ("class:item.sel", name),
                      ("",              "   "),
                      ("class:tag",     tag + "\n")]
        else:
            lines += [("",             "    "),
                      ("class:item",   name),
                      ("",             "   "),
                      ("class:tag",    tag + "\n")]

    lines.append(("", "\n"))
    lines.append(("class:hint",    "  "))
    lines.append(("class:hint-key","↑ ↓"))
    lines.append(("class:hint",    " navigate   "))
    lines.append(("class:hint-key","enter"))
    lines.append(("class:hint",    " select   "))
    lines.append(("class:hint-key","q"))
    lines.append(("class:hint",    " quit\n"))
    return lines

def run_menu(items: list, title: str = "SELECT NETWORK", show_art: bool = True) -> int | None:
    """
    Generic arrow-key menu. Returns selected index or None if quit.
    """
    selected = [0]
    chosen   = [None]
    kb       = KeyBindings()

    @kb.add("up")
    def _up(event):
        selected[0] = (selected[0] - 1) % len(items)
        event.app.invalidate()

    @kb.add("down")
    def _down(event):
        selected[0] = (selected[0] + 1) % len(items)
        event.app.invalidate()

    @kb.add("enter")
    def _enter(event):
        chosen[0] = selected[0]
        event.app.exit()

    @kb.add("q")
    @kb.add("c-c")
    def _quit(event):
        event.app.exit()

    layout = Layout(HSplit([
        Window(content=FormattedTextControl(
            lambda: build_menu(selected[0], items, title, show_art)
        ), dont_extend_height=True)
    ]))

    app = Application(layout=layout, key_bindings=kb, style=STYLE,
                      full_screen=False, mouse_support=False)
    app.run()
    return chosen[0]

def select_network() -> str | None:
    idx = run_menu(NETWORKS, "SELECT NETWORK")
    if idx is None:
        return None
    network = NETWORKS[idx][0].lower()

    # evm — show chain submenu
    if network == "eth":
        evm_idx = run_menu(EVM_NETWORKS, "SELECT EVM NETWORK", show_art=False)
        if evm_idx is None:
            return None
        return EVM_NETWORKS[evm_idx][0].lower()

    return network

def ask_params(network: str) -> dict | None:
    clear()
    print(f"\n  TXFetch  ›  {network.upper()}\n")
    print("  " + "─" * 48)

    print("\n  [1/4]  Date (UTC)")
    print("         Single:  07.04.2026")
    print("         Range:   07.04.2026-09.04.2026")
    date_range_mode = False
    time_from = time_to = None
    while True:
        raw_date = prompt("         › ").strip()
        # try date range first (two dates separated by dash)
        if raw_date.count(".") >= 4:
            mid = raw_date.find(".", 6)  # find dot after first date
            sep_idx = raw_date.find("-", mid) if mid > 0 else -1
            if sep_idx > 0:
                d1 = _parse_date(raw_date[:sep_idx].strip())
                d2 = _parse_date(raw_date[sep_idx+1:].strip())
                if d1 and d2:
                    time_from = datetime(d1.year, d1.month, d1.day, 0, 0, 0, tzinfo=timezone.utc)
                    time_to   = datetime(d2.year, d2.month, d2.day, 23, 59, 59, tzinfo=timezone.utc)
                    date_range_mode = True
                    break
        date = _parse_date(raw_date)
        if date is None:
            print("         Could not parse. Examples: 07.04.2026  or  07.04.2026-09.04.2026")
            continue
        break

    if not date_range_mode:
        print(f"\n  [2/4]  Time (UTC)  —  date: {date.strftime('%d.%m.%Y')}")
        print("         Format: HH:MM   or   HH:MM:SS")
        while True:
            raw_time = prompt("         › ").strip()
            time_result = _parse_time(raw_time)
            if time_result is None:
                print("         Could not parse time. Examples: 17:38  or  17:38:45")
                continue
            t, has_seconds = time_result
            break

        anchor = datetime(
            date.year, date.month, date.day,
            t.hour, t.minute, t.second,
            tzinfo=timezone.utc
        )

        clear()
        print(f"\n  TXFetch  ›  {network.upper()}\n")
        print("  " + "─" * 48)
        print(f"\n  Anchor time: {anchor.strftime('%d.%m.%Y %H:%M:%S')} UTC\n")

        range_items = [(n, d) for n, d, _ in TIME_RANGES]
        range_idx = run_menu(range_items, "SELECT TIME RANGE", show_art=False)

        if range_idx is None:
            return None

        range_key = TIME_RANGES[range_idx][2]

        if range_key == "minute":
            time_from = anchor.replace(second=0)
            time_to   = anchor.replace(second=59)
        elif range_key == "second":
            time_from = anchor
            time_to   = anchor
        elif range_key == "1min":
            time_from = anchor - timedelta(minutes=1)
            time_to   = anchor + timedelta(minutes=1)
        elif range_key == "5min":
            time_from = anchor - timedelta(minutes=5)
            time_to   = anchor + timedelta(minutes=5)
        elif range_key == "15min":
            time_from = anchor - timedelta(minutes=15)
            time_to   = anchor + timedelta(minutes=15)
        elif range_key == "custom":
            clear()
            print(f"\n  TXFetch  ›  {network.upper()}  ›  Custom range\n")
            print("  " + "─" * 48)
            print(f"\n  Anchor: {anchor.strftime('%d.%m.%Y %H:%M:%S')} UTC\n")
            print("  Enter start time (HH:MM or HH:MM:SS):")
            while True:
                r = _parse_time(prompt("  Start › ").strip())
                if r is None:
                    print("  Invalid format.")
                    continue
                ts, _ = r
                time_from = anchor.replace(hour=ts.hour, minute=ts.minute, second=ts.second)
                break
            print("  Enter end time (HH:MM or HH:MM:SS):")
            while True:
                r = _parse_time(prompt("  End   › ").strip())
                if r is None:
                    print("  Invalid format.")
                    continue
                te, _ = r
                time_to = anchor.replace(hour=te.hour, minute=te.minute, second=te.second)
                break
    else:
        print(f"\n  [2/4]  Time — skipped (full date range)")

    clear()
    print(f"\n  TXFetch  ›  {network.upper()}\n")
    print("  " + "─" * 48)
    _show_window(time_from, time_to)

    coins = COINS.get(network, [])
    print(f"\n  [3/4]  Coin")
    for i, c in enumerate(coins, 1):
        print(f"         {i}. {c}")
    while True:
        raw_coin = prompt("         › ").strip().upper()
        if raw_coin in coins:
            coin = raw_coin
            break
        try:
            idx = int(raw_coin) - 1
            if 0 <= idx < len(coins):
                coin = coins[idx]
                break
        except ValueError:
            pass
        print(f"         Enter name or number from the list.")

    print(f"\n  [4/4]  Amount in {coin}  (Enter to skip)")
    print("         Single value: 800   Range: 500-800")
    raw_amount = prompt("         › ").strip()
    amount    = None
    amount_to = None
    if raw_amount:
        if "-" in raw_amount:
            parts = raw_amount.split("-", 1)
            try:
                amount    = float(parts[0].strip().replace(",", "."))
                amount_to = float(parts[1].strip().replace(",", "."))
                print(f"         Range: {amount} — {amount_to}")
            except ValueError:
                print("         Could not parse range — skipping.")
                amount = amount_to = None
        else:
            try:
                amount = float(raw_amount.replace(",", "."))
            except ValueError:
                print("         Not a number — skipping.")

    print(f"\n  [5/5]  Memo / comment  (Enter to skip)")
    memo = prompt("         › ").strip() or None

    print("\n  " + "─" * 48)
    print(f"  Network:  {network.upper()}")
    print(f"  Coin:     {coin}")
    _show_window(time_from, time_to)
    if amount is not None and amount_to is not None:
        print(f"  Amount:   {amount} — {amount_to}")
    else:
        print(f"  Amount:   {amount if amount is not None else '—'}")
    print(f"  Memo:     {memo if memo else '—'}")
    print("  " + "─" * 48)

    confirm = prompt("\n  Start search? [Y/n] › ").strip().lower()
    if confirm in ("n", "no"):
        return None

    return {
        "network":   network,
        "coin":      coin,
        "time_from": time_from,
        "time_to":   time_to,
        "amount":    amount,
        "amount_to": amount_to,
        "memo":      memo,
    }

def _show_window(time_from: datetime, time_to: datetime):
    """prints the search window in a clear, visible format."""
    tf = time_from.strftime("%d.%m.%Y %H:%M:%S")
    tt = time_to.strftime("%H:%M:%S")
    delta = int((time_to - time_from).total_seconds())
    print(f"  Window:   {tf} → {tt} UTC  ({delta}s)")

def _parse_date(s: str) -> datetime | None:
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def _parse_time(s: str) -> tuple[datetime, bool] | None:
    for fmt, has_sec in (("%H:%M:%S", True), ("%H:%M", False)):
        try:
            dt = datetime.strptime(s, fmt)
            return dt, has_sec
        except ValueError:
            continue
    return None

def run_search(params: dict):
    from modules.models import TXQuery

    network = params["network"]
    query = TXQuery(
        time_from = params["time_from"],
        time_to   = params["time_to"],
        coin      = params["coin"],
        amount    = params.get("amount"),
        amount_to = params.get("amount_to"),
        memo      = params.get("memo"),
    )

    clear()
    print(f"\n  Searching...  [{network.upper()} / {params['coin']}]")
    _show_window(params["time_from"], params["time_to"])
    print()

    def _run_adapter(q):
        if network == "tron":
            from modules.tron import TronAdapter
            return TronAdapter().fetch_candidates(q)
        elif network == "ton":
            from modules.ton import TonAdapter
            return TonAdapter().fetch_candidates(q)
        elif network == "eth":
            from modules.eth import ETHAdapter
            return ETHAdapter().fetch_candidates(q)
        elif network == "bsc":
            from modules.bsc import BSCAdapter
            return BSCAdapter().fetch_candidates(q)
        elif network == "base":
            from modules.base_chain import BaseAdapter
            return BaseAdapter().fetch_candidates(q)
        elif network == "btc":
            from modules.btc import BTCAdapter
            return BTCAdapter().fetch_candidates(q)
        else:
            print("  unknown network.")
            return []

    try:
        results = _run_adapter(query)

        # if memo was specified but nothing found — retry without it
        if not results and query.memo:
            print(f"\n  no matches with memo '{query.memo}' — retrying without it...")
            from dataclasses import replace
            query_no_memo = replace(query, memo=None)
            results = _run_adapter(query_no_memo)
            if results:
                print("  found candidates without memo — memo scoring skipped.\n")

    except Exception as e:
        print(f"  error: {e}")
        return

    _print_results(results)

def _print_results(results):
    if not results:
        print("  No candidates found.\n")
        return

    count = len(results)
    print(f"  Found: {count} candidate(s)\n")

    if count > EXCEL_THRESHOLD:
        print(f"  Large result set ({count}). Enter a limit or press Enter to export all to Excel.")
        raw_limit = prompt("  Limit (or Enter → Excel) › ").strip()
        if not raw_limit:
            _export_excel(results)
            return
        try:
            results = results[:int(raw_limit)]
        except ValueError:
            print("  Not a number — showing first 50.")
            results = results[:50]

    _print_table(results)
    _copy_txid(results)

    save = prompt("  Save to Excel? [y/N] › ").strip().lower()
    if save in ("y", "yes"):
        _export_excel(results)

def _print_table(results):
    print("  " + "─" * 62)
    print(f"  {'#':<4} {'TXID (short)':<24} {'Conf':>6}  {'Amount':>12}  Time")
    print("  " + "─" * 62)
    for i, c in enumerate(results, 1):
        tx = c.tx_id
        txid_short = tx[:20] + "…" if len(tx) > 20 else tx
        ts   = c.timestamp.strftime("%H:%M:%S")
        amt  = f"{c.amount:.2f}"
        conf = f"{c.confidence:.0%}"
        print(f"  {i:<4} {txid_short:<24} {conf:>6}  {amt:>12}  {ts}")
    print("  " + "─" * 62)

def _copy_txid(results):
    """ask user to pick a result number and print the full TXID."""
    print()
    raw = prompt("  Enter # to get full TXID  (or Enter to skip) › ").strip()
    if not raw:
        return
    try:
        idx = int(raw) - 1
        if 0 <= idx < len(results):
            print(f"\n  Full TXID #{idx+1}:")
            print(f"  {results[idx].tx_id}\n")
        else:
            print("  Number out of range.")
    except ValueError:
        print("  Invalid input.")

def _export_excel(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TXFetch Results"

    headers = ["#", "TX ID", "Network", "Coin", "Amount", "Time (UTC)",
               "From", "To", "Memo", "Confidence"]
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="5a9f7a", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, c in enumerate(results, 1):
        ws.append([
            i,
            c.tx_id,
            c.network,
            c.coin,
            c.amount,
            c.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            c.from_address,
            c.to_address,
            c.memo or "",
            f"{c.confidence:.0%}",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    filename = f"txfetch_{results[0].network.lower()}_{results[0].timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"\n  Saved: {filename}\n")

def _parse_args():
    parser = argparse.ArgumentParser(
        prog="txfetch",
        description="blockchain transaction retrieval tool",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "examples:\n"
            "  txf --network tron --coin USDT --date 07.04.2026 --time 17:38 --amount 800\n"
            "  txf --network ton  --coin TON  --date 14.04.2026 --time 23:39\n"
            "  txf --network btc  --coin BTC  --date 02.05.2026 --time 00:24 --amount 0.5\n"
        )
    )
    parser.add_argument("--network", "-n",
        choices=["tron", "ton", "eth", "bsc", "base", "btc"],
        help="network to search on")
    parser.add_argument("--coin", "-c",
        help="coin to search (e.g. USDT, BTC, TON)")
    parser.add_argument("--date", "-d",
        help="date in UTC — DD.MM.YYYY or YYYY-MM-DD")
    parser.add_argument("--time", "-t",
        help="time in UTC — HH:MM or HH:MM:SS")
    parser.add_argument("--amount", "-a",
        type=float, default=None,
        help="expected amount (optional)")
    parser.add_argument("--memo", "-m",
        default=None,
        help="memo / comment (optional)")
    parser.add_argument("--range", "-r",
        choices=["second", "minute", "1min", "5min", "15min"],
        default="minute",
        help=(
            "time window around the given timestamp (default: minute)\n"
            "  second — exact second\n"
            "  minute — full minute HH:MM:00 to HH:MM:59\n"
            "  1min   — ±1 minute\n"
            "  5min   — ±5 minutes\n"
            "  15min  — ±15 minutes"
        ))
    parser.add_argument("--tolerance",
        type=float, default=0.01,
        help="amount tolerance as fraction, e.g. 0.01 = ±1%% (default: 0.01)")
    return parser.parse_args()


def _build_time_window(anchor: datetime, range_key: str):
    """build time_from / time_to from anchor and range key."""
    if range_key == "second":
        return anchor, anchor
    elif range_key == "minute":
        return anchor.replace(second=0), anchor.replace(second=59)
    elif range_key == "1min":
        return anchor - timedelta(minutes=1), anchor + timedelta(minutes=1)
    elif range_key == "5min":
        return anchor - timedelta(minutes=5), anchor + timedelta(minutes=5)
    elif range_key == "15min":
        return anchor - timedelta(minutes=15), anchor + timedelta(minutes=15)
    return anchor.replace(second=0), anchor.replace(second=59)


def main():
    args = _parse_args()

    # if all required args provided — run directly without TUI
    if args.network and args.coin and args.date and args.time:
        date = _parse_date(args.date)
        if date is None:
            print(f"  error: could not parse date '{args.date}'")
            sys.exit(1)

        time_result = _parse_time(args.time)
        if time_result is None:
            print(f"  error: could not parse time '{args.time}'")
            sys.exit(1)

        t, _ = time_result
        anchor = datetime(
            date.year, date.month, date.day,
            t.hour, t.minute, t.second,
            tzinfo=timezone.utc
        )
        time_from, time_to = _build_time_window(anchor, args.range)

        params = {
            "network":   args.network,
            "coin":      args.coin.upper(),
            "time_from": time_from,
            "time_to":   time_to,
            "amount":    args.amount,
            "memo":      args.memo,
        }

        os.system("cls" if os.name == "nt" else "clear")
        run_search(params)
        sys.exit(0)

    # no args — launch TUI
    os.system("cls" if os.name == "nt" else "clear")

    while True:
        network = select_network()

        if network is None:
            clear()
            print("\n  Goodbye.\n")
            sys.exit(0)

        params = ask_params(network)
        if params is None:
            continue

        run_search(params)

        again = prompt("\n  New search? [Y/n] › ").strip().lower()
        if again in ("n", "no"):
            clear()
            print("\n  Goodbye.\n")
            sys.exit(0)

if __name__ == "__main__":
    main()
